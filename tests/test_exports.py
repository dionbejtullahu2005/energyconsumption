from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from enerco_analysis.exports import build_cluster_company_export


def test_cluster_company_export_has_one_sheet_per_group() -> None:
    assignments = pd.DataFrame({
        "company_id": ["Kompania 1", "Kompania 2", "Kompania 3"],
        "cluster_number": [1, 1, 2],
        "cluster_id": ["Klaster 1", "Klaster 1", "Klaster 2"],
        "cluster_description": ["Profili A", "Profili A", "Profili B"],
    })
    energy = pd.DataFrame({
        "company_id": ["Kompania 1", "Kompania 2", "Kompania 3"],
        "energy_total_kwh": [1000.0, 2500.0, 4000.0],
    })
    payload = build_cluster_company_export(
        assignments, energy, pd.Timestamp("2025-06-01").date(), pd.Timestamp("2025-06-30").date()
    )
    workbook = load_workbook(BytesIO(payload), data_only=False)
    assert workbook.sheetnames == ["Përmbledhje", "Grupi 1", "Grupi 2"]
    assert workbook["Grupi 1"].max_row == 3
    assert workbook["Grupi 2"]["A2"].value == "Kompania 3"
    assert workbook["Përmbledhje"]["C2"].value == "=COUNTA('Grupi 1'!A2:A3)"
    assert workbook["Përmbledhje"]["D3"].value == "=SUM('Grupi 2'!D2:D2)"
