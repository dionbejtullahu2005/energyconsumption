from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta
from math import sqrt
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook

TIME_COLUMNS = ("PeriodYear", "PeriodMonth", "WeekDay", "Date", "Hour", "Tariff")

@dataclass
class SeriesStats:
    sheet: str
    company_id: str
    meter_id: str
    energy_flow: str
    expected_hours: int = 0
    missing: int = 0
    negative: int = 0
    non_numeric: int = 0
    zero_hours: int = 0
    zero_run_events: int = 0
    zero_run_hours: int = 0
    maximum_zero_run: int = 0
    current_zero_run: int = 0
    current_zero_start: tuple[Any, Any] | None = None
    count: int = 0
    total: float = 0.0
    total_squares: float = 0.0
    minimum: float | None = None
    maximum: float | None = None
    extreme_count: int = 0
    first_valid_position: int | None = None
    last_valid_position: int | None = None
    zero_run_details: list[dict[str, Any]] = field(default_factory=list)

    def add_value(
        self,
        value: Any,
        row_key: tuple[Any, Any],
        position: int,
        zero_run_threshold: int,
        anomalies: list[dict[str, Any]],
    ) -> None:
        if value is None or (isinstance(value, str) and not value.strip()):
            self.missing += 1
            self._close_zero_run(row_key, zero_run_threshold)
            return
        try:
            number = float(value)
        except (TypeError, ValueError):
            self.non_numeric += 1
            self._close_zero_run(row_key, zero_run_threshold)
            anomalies.append(self._anomaly(row_key, value, "Vlerë jonumerike"))
            return

        self.count += 1
        if self.first_valid_position is None:
            self.first_valid_position = position
        self.last_valid_position = position
        self.total += number
        self.total_squares += number * number
        self.minimum = number if self.minimum is None else min(self.minimum, number)
        self.maximum = number if self.maximum is None else max(self.maximum, number)
        if number < 0:
            self.negative += 1
            anomalies.append(self._anomaly(row_key, number, "Vlerë negative"))
        if number == 0:
            self.zero_hours += 1
            if self.current_zero_run == 0:
                self.current_zero_start = row_key
            self.current_zero_run += 1
        else:
            self._close_zero_run(row_key, zero_run_threshold)

    def finish(self, zero_run_threshold: int) -> None:
        self._close_zero_run(None, zero_run_threshold)

    def _close_zero_run(
        self, next_key: tuple[Any, Any] | None, zero_run_threshold: int
    ) -> None:
        if self.current_zero_run >= zero_run_threshold:
            self.zero_run_events += 1
            self.zero_run_hours += self.current_zero_run
            self.zero_run_details.append(
                {
                    "sheet": self.sheet,
                    "company_id": self.company_id,
                    "meter_id": self.meter_id,
                    "energy_flow": self.energy_flow,
                    "start_date": self.current_zero_start[0] if self.current_zero_start else None,
                    "start_hour": self.current_zero_start[1] if self.current_zero_start else None,
                    "run_hours": self.current_zero_run,
                    "next_date": next_key[0] if next_key else None,
                    "next_hour": next_key[1] if next_key else None,
                }
            )
        self.maximum_zero_run = max(self.maximum_zero_run, self.current_zero_run)
        self.current_zero_run = 0
        self.current_zero_start = None

    def _anomaly(self, row_key: tuple[Any, Any], value: Any, reason: str) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "company_id": self.company_id,
            "meter_id": self.meter_id,
            "energy_flow": self.energy_flow,
            "date": row_key[0],
            "hour": row_key[1],
            "value": value,
            "reason": reason,
        }

    @property
    def mean(self) -> float | None:
        return self.total / self.count if self.count else None

    @property
    def standard_deviation(self) -> float | None:
        if self.count < 2:
            return None
        variance = max(0.0, (self.total_squares - self.total**2 / self.count) / (self.count - 1))
        return sqrt(variance)

    @property
    def leading_missing(self) -> int:
        return self.first_valid_position if self.first_valid_position is not None else self.expected_hours

    @property
    def trailing_missing(self) -> int:
        if self.last_valid_position is None:
            return 0
        return max(0, self.expected_hours - self.last_valid_position - 1)

    @property
    def active_span_hours(self) -> int:
        if self.first_valid_position is None or self.last_valid_position is None:
            return 0
        return self.last_valid_position - self.first_valid_position + 1

    @property
    def internal_missing(self) -> int:
        return max(0, self.active_span_hours - self.count - self.non_numeric)


def _parse_meter_header(raw_header: Any, is_prosumer: bool) -> tuple[str, str]:
    header = str(raw_header).strip()
    if is_prosumer and " - " in header:
        meter_id, suffix = header.rsplit(" - ", 1)
        if suffix in {"A+", "A-"}:
            return meter_id, "consumption_import" if suffix == "A+" else "injection_export"
    return header, "consumption_import"


def _expected_time_keys(start: str, end: str) -> list[tuple[Any, int]]:
    start_date = datetime.fromisoformat(start).date()
    end_date = datetime.fromisoformat(end).date()
    keys: list[tuple[Any, int]] = []
    current = start_date
    while current <= end_date:
        keys.extend((current, hour) for hour in range(1, 25))
        current += timedelta(days=1)
    return keys


def _normalise_date(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "date"):
        return value.date()
    return value


def build_quality_report(
    workbook_path: Path,
    output_path: Path,
    config: dict[str, Any],
    machine_table_path: Path | None = None,
) -> dict[str, Any]:
    quality_config = config["quality"]
    zero_threshold = int(quality_config["zero_run_hours"])
    expected_keys = _expected_time_keys(
        config["analysis"]["start_date"], config["analysis"]["end_date"]
    )
    expected_set = set(expected_keys)

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    all_stats: list[SeriesStats] = []
    anomalies: list[dict[str, Any]] = []
    timeline_rows: list[dict[str, Any]] = []
    sheet_rows: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        row_iter = sheet.iter_rows(values_only=True)
        company_headers = next(row_iter)
        meter_headers = next(row_iter)
        is_prosumer = config["inputs"]["prosumer_sheet_match"].lower() in sheet_name.lower()
        stats: list[SeriesStats] = []
        for company, raw_meter in zip(company_headers[6:], meter_headers[6:]):
            meter_id, flow = _parse_meter_header(raw_meter, is_prosumer)
            stats.append(SeriesStats(sheet_name, str(company).strip(), meter_id, flow))

        actual_keys: list[tuple[Any, int]] = []
        metadata_mismatches = 0
        for position, values in enumerate(row_iter):
            date_value = _normalise_date(values[3])
            try:
                hour_value = int(values[4])
            except (TypeError, ValueError):
                hour_value = values[4]
            row_key = (date_value, hour_value)
            actual_keys.append(row_key)
            if date_value is not None:
                if values[0] != date_value.year or values[1] != date_value.month:
                    metadata_mismatches += 1
            for item, value in zip(stats, values[6:]):
                item.add_value(value, row_key, position, zero_threshold, anomalies)

        for item in stats:
            item.expected_hours = len(expected_keys)
            item.finish(zero_threshold)
        all_stats.extend(stats)

        actual_set = set(actual_keys)
        missing_keys = expected_set.difference(actual_set)
        extra_keys = actual_set.difference(expected_set)
        duplicate_count = len(actual_keys) - len(actual_set)
        invalid_hours = sum(not isinstance(h, int) or not 1 <= h <= 24 for _, h in actual_keys)
        sheet_rows.append(
            {
                "sheet": sheet_name,
                "data_rows": len(actual_keys),
                "series_columns": len(stats),
                "physical_meters": len({s.meter_id for s in stats}),
                "missing_time_keys": len(missing_keys),
                "extra_time_keys": len(extra_keys),
                "duplicate_time_keys": duplicate_count,
                "invalid_hours": invalid_hours,
                "date_metadata_mismatches": metadata_mismatches,
            }
        )
        for key in sorted(missing_keys, key=str):
            timeline_rows.append({"sheet": sheet_name, "date": key[0], "hour": key[1], "issue": "Mungon"})
        for key in sorted(extra_keys, key=str):
            timeline_rows.append({"sheet": sheet_name, "date": key[0], "hour": key[1], "issue": "Jashtë periudhës"})

    workbook.close()

    # Kalimi i dytë numëron vlerat 50× mbi mesataren historike të serisë.
    stat_lookup = {(s.sheet, s.meter_id, s.energy_flow): s for s in all_stats}
    extreme_multiple = float(quality_config["extreme_multiple_of_mean"])
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        row_iter = sheet.iter_rows(values_only=True)
        next(row_iter)
        meter_headers = next(row_iter)
        is_prosumer = config["inputs"]["prosumer_sheet_match"].lower() in sheet_name.lower()
        identities = [_parse_meter_header(value, is_prosumer) for value in meter_headers[6:]]
        for values in row_iter:
            row_key = (_normalise_date(values[3]), values[4])
            for identity, value in zip(identities, values[6:]):
                stat = stat_lookup[(sheet_name, identity[0], identity[1])]
                if stat.mean is None or stat.mean <= 0:
                    continue
                try:
                    number = float(value)
                except (TypeError, ValueError):
                    continue
                if number > extreme_multiple * stat.mean:
                    stat.extreme_count += 1
                    anomalies.append(stat._anomaly(row_key, number, f"> {extreme_multiple:g}× mesatarja"))
    workbook.close()

    unusable_threshold = float(quality_config["unusable_missing_share"])
    report_rows: list[dict[str, Any]] = []
    zero_runs = [detail for stat in all_stats for detail in stat.zero_run_details]
    for stat in all_stats:
        missing_share = stat.missing / stat.expected_hours if stat.expected_hours else 1.0
        internal_missing_share = (
            stat.internal_missing / stat.active_span_hours if stat.active_span_hours else 1.0
        )
        if missing_share > unusable_threshold:
            status = "I papërdorshëm"
        elif any((stat.missing, stat.negative, stat.non_numeric, stat.zero_run_events, stat.extreme_count)):
            status = "Për shqyrtim"
        else:
            status = "I pastër"
        if stat.active_span_hours == 0 or internal_missing_share > unusable_threshold:
            active_status = "I papërdorshëm"
        elif any((stat.internal_missing, stat.negative, stat.non_numeric, stat.zero_run_events, stat.extreme_count)):
            active_status = "Për shqyrtim"
        else:
            active_status = "I pastër"
        if stat.first_valid_position is None:
            missing_pattern = "Pa lexime"
        elif stat.leading_missing and stat.trailing_missing:
            missing_pattern = "Fillim i vonë dhe fund i hershëm"
        elif stat.leading_missing:
            missing_pattern = "Fillim i vonë"
        elif stat.trailing_missing:
            missing_pattern = "Fund i hershëm"
        elif stat.internal_missing:
            missing_pattern = "Boshllëqe të brendshme"
        else:
            missing_pattern = "Pa mungesa"
        report_rows.append(
            {
                "sheet": stat.sheet,
                "company_id": stat.company_id,
                "meter_id": stat.meter_id,
                "energy_flow": stat.energy_flow,
                "expected_hours": stat.expected_hours,
                "valid_numeric_hours": stat.count,
                "missing_hours": stat.missing,
                "missing_share": missing_share,
                "leading_missing_hours": stat.leading_missing,
                "internal_missing_hours": stat.internal_missing,
                "trailing_missing_hours": stat.trailing_missing,
                "active_span_hours": stat.active_span_hours,
                "internal_missing_share_active_span": internal_missing_share,
                "missing_pattern": missing_pattern,
                "negative_values": stat.negative,
                "non_numeric_values": stat.non_numeric,
                "zero_hours": stat.zero_hours,
                "zero_run_events_48h_plus": stat.zero_run_events,
                "zero_run_hours_48h_plus": stat.zero_run_hours,
                "maximum_zero_run_hours": stat.maximum_zero_run,
                "extreme_values_50x_mean": stat.extreme_count,
                "mean_kwh": stat.mean,
                "std_kwh": stat.standard_deviation,
                "min_kwh": stat.minimum,
                "max_kwh": stat.maximum,
                "quality_status": status,
                "active_period_quality_status": active_status,
            }
        )

    report_df = pd.DataFrame(report_rows)
    summary = {
        "workbook": workbook_path.name,
        "sheet_count": len(sheet_rows),
        "series_count": len(report_df),
        "physical_meter_count": int(report_df["meter_id"].nunique()),
        "company_count": int(report_df["company_id"].nunique()),
        "clean_series": int((report_df["quality_status"] == "I pastër").sum()),
        "review_series": int((report_df["quality_status"] == "Për shqyrtim").sum()),
        "unusable_series": int((report_df["quality_status"] == "I papërdorshëm").sum()),
        "active_period_clean_series": int(
            (report_df["active_period_quality_status"] == "I pastër").sum()
        ),
        "active_period_review_series": int(
            (report_df["active_period_quality_status"] == "Për shqyrtim").sum()
        ),
        "active_period_unusable_series": int(
            (report_df["active_period_quality_status"] == "I papërdorshëm").sum()
        ),
        "timeline_issue_count": len(timeline_rows),
        "anomaly_count": len(anomalies),
    }
    summary_df = pd.DataFrame({"metric": summary.keys(), "value": summary.values()})
    if machine_table_path is not None:
        machine_table_path.parent.mkdir(parents=True, exist_ok=True)
        report_df.to_parquet(machine_table_path, index=False)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        pd.DataFrame(sheet_rows).to_excel(writer, sheet_name="Sheet checks", index=False)
        report_df.to_excel(writer, sheet_name="Meter quality", index=False)
        pd.DataFrame(zero_runs).to_excel(writer, sheet_name="Zero runs", index=False)
        pd.DataFrame(anomalies).to_excel(writer, sheet_name="Value anomalies", index=False)
        pd.DataFrame(timeline_rows).to_excel(writer, sheet_name="Timeline issues", index=False)
    return summary
