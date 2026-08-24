from __future__ import annotations

from datetime import date
from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_sheet(sheet, widths: list[float]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def build_cluster_company_export(
    assignments: pd.DataFrame,
    period_company_energy: pd.DataFrame,
    period_start: date,
    period_end: date,
) -> bytes:
    """Krijon workbook anonim me një fletë përmbledhëse dhe një fletë për grup."""
    required = {"company_id", "cluster_number", "cluster_id", "cluster_description"}
    missing = sorted(required.difference(assignments.columns))
    if missing:
        raise ValueError(f"Mungojnë kolonat e klasterëve për eksport: {missing}")
    if not {"company_id", "energy_total_kwh"}.issubset(period_company_energy.columns):
        raise ValueError("Mungojnë company_id/energy_total_kwh për eksport")

    details = assignments[list(required)].drop_duplicates("company_id").merge(
        period_company_energy[["company_id", "energy_total_kwh"]],
        on="company_id", how="left", validate="one_to_one",
    )
    details["energy_total_kwh"] = details["energy_total_kwh"].fillna(0.0)
    details["energy_total_mwh"] = details["energy_total_kwh"] / 1000
    details = details.sort_values(["cluster_number", "company_id"])

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Përmbledhje"
    summary.append([
        "Grupi", "Përshkrimi i grupit", "Numri i kompanive",
        "Energjia totale (kWh)", "Energjia totale (MWh)", "Data fillestare", "Data përfundimtare",
    ])

    for row_number, (cluster_number, group) in enumerate(
        details.groupby("cluster_number", observed=True), start=2
    ):
        sheet_name = f"Grupi {int(cluster_number)}"
        sheet = workbook.create_sheet(sheet_name)
        sheet.append([
            "Kompania", "Grupi", "Përshkrimi i grupit",
            "Energjia totale (kWh)", "Energjia totale (MWh)",
            "Data fillestare", "Data përfundimtare",
        ])
        for item in group.itertuples(index=False):
            sheet.append([
                item.company_id, sheet_name, item.cluster_description,
                float(item.energy_total_kwh), float(item.energy_total_mwh),
                period_start, period_end,
            ])
        last_row = sheet.max_row
        summary.append([
            sheet_name, str(group.iloc[0]["cluster_description"]),
            f"=COUNTA('{sheet_name}'!A2:A{last_row})",
            f"=SUM('{sheet_name}'!D2:D{last_row})",
            f"=D{row_number}/1000", period_start, period_end,
        ])
        for cell in sheet["D"][1:]:
            cell.number_format = "#,##0.00"
        for cell in sheet["E"][1:]:
            cell.number_format = "#,##0.00"
        for column in ("F", "G"):
            for cell in sheet[column][1:]:
                cell.number_format = "yyyy-mm-dd"
        _style_sheet(sheet, [18, 12, 70, 24, 24, 16, 18])

    for column in ("D", "E"):
        for cell in summary[column][1:]:
            cell.number_format = "#,##0.00"
    for column in ("F", "G"):
        for cell in summary[column][1:]:
            cell.number_format = "yyyy-mm-dd"
    _style_sheet(summary, [12, 70, 22, 24, 24, 16, 18])

    buffer = BytesIO()
    workbook.save(buffer)
    # Rihapja garanton që bajtet e prodhuara janë workbook i vlefshëm.
    load_workbook(BytesIO(buffer.getvalue()), read_only=True, data_only=False).close()
    return buffer.getvalue()
