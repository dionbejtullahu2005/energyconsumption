from __future__ import annotations

import json
import os
from collections import Counter
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

from enerco_analysis.quality import _normalise_date, _parse_meter_header


def _chunks(iterator: Any, size: int) -> Any:
    batch: list[tuple[Any, ...]] = []
    for row in iterator:
        batch.append(row)
        if len(batch) == size:
            yield batch
            batch = []
    if batch:
        yield batch


def _quality_lookup(quality_report_path: Path) -> dict[tuple[str, str, str], dict[str, Any]]:
    if quality_report_path.suffix.lower() == ".parquet":
        frame = pd.read_parquet(quality_report_path)
    else:
        frame = pd.read_excel(quality_report_path, sheet_name="Meter quality")
    lookup: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in frame.to_dict("records"):
        key = (str(row["sheet"]), str(row["meter_id"]), str(row["energy_flow"]))
        lookup[key] = row
    return lookup


def _build_long_table(
    rows: list[tuple[Any, ...]],
    row_offset: int,
    sheet_name: str,
    companies: np.ndarray,
    meters: np.ndarray,
    flows: np.ndarray,
    quality_rows: list[dict[str, Any]],
    peak_start: int,
    peak_end: int,
) -> pa.Table:
    series_count = len(meters)
    time_count = len(rows)
    dates = np.array([_normalise_date(row[3]) for row in rows], dtype=object)
    hours = np.array([int(row[4]) for row in rows], dtype=np.int16)
    tariffs = np.array([str(row[5]) if row[5] is not None else None for row in rows], dtype=object)
    interval_starts = pd.to_datetime(dates) + pd.to_timedelta(hours - 1, unit="h")
    energy_matrix = np.array([row[6 : 6 + series_count] for row in rows], dtype=object)
    energy = pd.to_numeric(pd.Series(energy_matrix.reshape(-1)), errors="coerce").to_numpy()
    energy_missing = np.isnan(energy)

    positions = np.repeat(np.arange(row_offset, row_offset + time_count), series_count)
    leading = np.tile(np.array([int(q["leading_missing_hours"]) for q in quality_rows]), time_count)
    active_span = np.tile(np.array([int(q["active_span_hours"]) for q in quality_rows]), time_count)
    in_active_span = (positions >= leading) & (positions < leading + active_span)

    repeated_dates = np.repeat(dates, series_count)
    repeated_hours = np.repeat(hours, series_count)
    repeated_starts = np.repeat(interval_starts.to_numpy(), series_count)
    repeated_weekdays = np.array([value.weekday() for value in dates], dtype=np.int8)

    data = {
        "company_id": np.tile(companies, time_count),
        "meter_id": np.tile(meters, time_count),
        "energy_flow": np.tile(flows, time_count),
        "interval_start": repeated_starts,
        "interval_end": repeated_starts + np.timedelta64(1, "h"),
        "date": repeated_dates,
        "year": np.repeat(np.array([value.year for value in dates], dtype=np.int16), series_count),
        "month": np.repeat(np.array([value.month for value in dates], dtype=np.int8), series_count),
        "weekday_number": np.repeat(repeated_weekdays, series_count),
        "is_weekend": np.repeat(repeated_weekdays >= 5, series_count),
        "hour_1_24": repeated_hours,
        "is_peak_07_18": (repeated_hours >= peak_start) & (repeated_hours <= peak_end),
        "tariff": np.repeat(tariffs, series_count),
        "energy_kwh": pa.array(energy, mask=energy_missing, type=pa.float64()),
        "is_missing": energy_missing,
        "in_active_span": in_active_span,
        "quality_status": np.tile(
            np.array([str(q["quality_status"]) for q in quality_rows], dtype=object), time_count
        ),
        "active_period_quality_status": np.tile(
            np.array([str(q["active_period_quality_status"]) for q in quality_rows], dtype=object),
            time_count,
        ),
        "source_sheet": np.full(time_count * series_count, sheet_name, dtype=object),
    }
    return pa.Table.from_pydict(data)


def transform_to_long(
    workbook_path: Path,
    quality_report_path: Path,
    output_path: Path,
    validation_path: Path,
    config: dict[str, Any],
    chunk_rows: int = 500,
) -> dict[str, Any]:
    quality = _quality_lookup(quality_report_path)
    expected_missing = sum(int(row["missing_hours"]) for row in quality.values())
    peak_start, peak_end = config["analysis"]["peak_hours"]
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    temporary_path = output_path.with_suffix(".tmp.parquet")
    if temporary_path.exists():
        temporary_path.unlink()

    writer: pq.ParquetWriter | None = None
    total_rows = 0
    total_missing = 0
    source_energy_sum = 0.0
    sheet_summaries: list[dict[str, Any]] = []
    seen_series: set[tuple[str, str]] = set()
    prosumer_directions: dict[str, set[str]] = {}

    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            iterator = sheet.iter_rows(values_only=True)
            company_headers = next(iterator)
            meter_headers = next(iterator)
            is_prosumer = config["inputs"]["prosumer_sheet_match"].lower() in sheet_name.lower()
            parsed = [_parse_meter_header(value, is_prosumer) for value in meter_headers[6:]]
            companies = np.array([str(value).strip() for value in company_headers[6:]], dtype=object)
            meters = np.array([item[0] for item in parsed], dtype=object)
            flows = np.array([item[1] for item in parsed], dtype=object)
            quality_rows = [quality[(sheet_name, meter, flow)] for meter, flow in parsed]

            local_identities = list(zip(meters.tolist(), flows.tolist()))
            if len(local_identities) != len(set(local_identities)):
                raise ValueError(f"Seri të dubluara në header-in e fletës {sheet_name}")
            overlap = seen_series.intersection(local_identities)
            if overlap:
                raise ValueError(f"Seri të përsëritura ndërmjet fletëve: {sorted(overlap)[:5]}")
            seen_series.update(local_identities)
            if is_prosumer:
                for meter, flow in local_identities:
                    prosumer_directions.setdefault(meter, set()).add(flow)

            sheet_output_rows = 0
            sheet_missing = 0
            sheet_sum = 0.0
            time_keys: set[tuple[Any, int]] = set()
            row_offset = 0
            for rows in _chunks(iterator, chunk_rows):
                for row in rows:
                    key = (_normalise_date(row[3]), int(row[4]))
                    if key in time_keys:
                        raise ValueError(f"Datë/orë e dubluar në {sheet_name}: {key}")
                    time_keys.add(key)
                table = _build_long_table(
                    rows,
                    row_offset,
                    sheet_name,
                    companies,
                    meters,
                    flows,
                    quality_rows,
                    int(peak_start),
                    int(peak_end),
                )
                if writer is None:
                    output_path.parent.mkdir(parents=True, exist_ok=True)
                    writer = pq.ParquetWriter(
                        temporary_path,
                        table.schema,
                        compression="zstd",
                        use_dictionary=True,
                    )
                writer.write_table(table)
                energy_column = table.column("energy_kwh")
                chunk_missing = energy_column.null_count
                chunk_sum = float(pa.compute.sum(energy_column).as_py() or 0.0)
                sheet_output_rows += table.num_rows
                sheet_missing += chunk_missing
                sheet_sum += chunk_sum
                row_offset += len(rows)

            expected_sheet_rows = len(time_keys) * len(meters)
            if sheet_output_rows != expected_sheet_rows:
                raise ValueError(
                    f"Numri i rreshtave nuk përputhet në {sheet_name}: "
                    f"{sheet_output_rows} != {expected_sheet_rows}"
                )
            sheet_summaries.append(
                {
                    "sheet": sheet_name,
                    "time_rows": len(time_keys),
                    "series": len(meters),
                    "physical_meters": len(set(meters.tolist())),
                    "output_rows": sheet_output_rows,
                    "missing_energy_rows": sheet_missing,
                    "energy_sum_kwh": sheet_sum,
                }
            )
            total_rows += sheet_output_rows
            total_missing += sheet_missing
            source_energy_sum += sheet_sum
    finally:
        workbook.close()
        if writer is not None:
            writer.close()

    pair_issues = {
        meter: sorted(flows)
        for meter, flows in prosumer_directions.items()
        if flows != {"consumption_import", "injection_export"}
    }
    if pair_issues:
        raise ValueError(f"Çifte jo të plota Prosumer: {pair_issues}")

    metadata = pq.ParquetFile(temporary_path).metadata
    parquet_rows = metadata.num_rows
    if parquet_rows != total_rows:
        raise ValueError(f"Validimi Parquet dështoi: {parquet_rows} != {total_rows}")
    if total_missing != expected_missing:
        raise ValueError(
            f"Mungesat nuk përputhen me Hapin 1: {total_missing} != {expected_missing}"
        )
    os.replace(temporary_path, output_path)

    flow_counts = Counter(flow for _, flow in seen_series)
    summary = {
        "source_workbook": workbook_path.name,
        "output_file": output_path.name,
        "sheet_count": len(sheet_summaries),
        "series_count": len(seen_series),
        "physical_meter_count": len({meter for meter, _ in seen_series}),
        "consumption_import_series": flow_counts["consumption_import"],
        "injection_export_series": flow_counts["injection_export"],
        "prosumer_pair_issues": len(pair_issues),
        "output_rows": total_rows,
        "missing_energy_rows": total_missing,
        "quality_report_missing_hours": expected_missing,
        "non_missing_energy_rows": total_rows - total_missing,
        "energy_sum_kwh": source_energy_sum,
        "duplicate_series_headers": 0,
        "duplicate_time_keys": 0,
        "parquet_metadata_rows": parquet_rows,
        "validation_passed": True,
    }
    validation_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(validation_path, engine="openpyxl") as excel:
        pd.DataFrame({"metric": summary.keys(), "value": summary.values()}).to_excel(
            excel, sheet_name="Summary", index=False
        )
        pd.DataFrame(sheet_summaries).to_excel(excel, sheet_name="By sheet", index=False)
    json_path = validation_path.with_suffix(".json")
    json_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    return summary
